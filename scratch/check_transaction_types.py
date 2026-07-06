import openpyxl

file_path = r'C:\Users\ahmad\Downloads\testing maret\income maret 2026.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

headers = [str(cell.value or '').strip() for cell in ws[1]]
type_col = next((i for i, h in enumerate(headers) if 'jenis transaksi' in h.lower() or 'transaction type' in h.lower()), -1)

types = {}
for r in range(2, ws.max_row + 1):
    val = str(ws.cell(row=r, column=type_col+1).value or '').strip()
    types[val] = types.get(val, 0) + 1

print("Distinct transaction types and counts:")
for k, v in types.items():
    print(f"  {k}: {v}")
wb.close()
