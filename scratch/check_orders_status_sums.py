import openpyxl

# Load orders first to filter by March creation
orders_file = r'C:\Users\ahmad\Downloads\testing maret\pesanan maret 2026.xlsx'
wb_ord = openpyxl.load_workbook(orders_file)
ws_ord = wb_ord.active
headers_ord = [str(cell.value or '').strip().lower() for cell in ws_ord[1]]
ord_id_col = next((i for i, h in enumerate(headers_ord) if 'order id' in h or 'id pesanan' in h), -1)
status_col = next((i for i, h in enumerate(headers_ord) if 'status' in h), -1)
price_col = next((i for i, h in enumerate(headers_ord) if 'subtotal setelah diskon penjual' in h or 'harga setelah diskon' in h or 'omset' in h or 'subtotal' in h), -1)
print(f"Price column found: {headers_ord[price_col]}")

status_sums = {}
for r in range(2, ws_ord.max_row + 1):
    status = str(ws_ord.cell(row=r, column=status_col+1).value or '').strip()
    price = 0.0
    val = ws_ord.cell(row=r, column=price_col+1).value
    if val is not None:
        try:
            price = float(val)
        except ValueError:
            pass
    status_sums[status] = status_sums.get(status, 0.0) + price

print("Sums by status in orders file:")
for k, v in status_sums.items():
    print(f"  {k}: {v:,.2f}")
wb_ord.close()
