import openpyxl

file_path = r'C:\Users\ahmad\Downloads\testing maret\income maret 2026.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

headers = [str(cell.value or '').strip() for cell in ws[1]]

# Let's search for any cell containing 'batal' or 'cancel' or 'status'
found = False
for r in range(2, min(500, ws.max_row + 1)):
    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(row=r, column=c).value or '')
        if 'batal' in val.lower() or 'cancel' in val.lower() or 'refund' in val.lower() or 'retur' in val.lower():
            print(f"  Row {r}, Col {c} ({headers[c-1]}): {val}")
            found = True
if not found:
    print("No cancellation/return status values found in cell content.")
wb.close()
