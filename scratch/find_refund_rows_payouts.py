import openpyxl

file_path = r'C:\Users\ahmad\Downloads\testing maret\income maret 2026.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

headers = [str(cell.value or '').strip() for cell in ws[1]]
type_col = next((i for i, h in enumerate(headers) if 'jenis transaksi' in h.lower() or 'transaction type' in h.lower()), -1)
refund_col = next((i for i, h in enumerate(headers) if 'pengembalian dana' in h.lower() or 'refund' in h.lower()), -1)
order_id_col = next((i for i, h in enumerate(headers) if 'pesanan' in h.lower() or 'order' in h.lower()), -1)

print("Rows in Keuangan file with non-zero refund column values:")
rows_found = 0
for r in range(2, ws.max_row + 1):
    ref_val = ws.cell(row=r, column=refund_col+1).value
    if ref_val is not None:
        try:
            val = float(ref_val)
            if abs(val) > 0.01:
                rows_found += 1
                if rows_found <= 20:
                    print(f"  Row {r}: OrderID={ws.cell(row=r, column=order_id_col+1).value}, Type={ws.cell(row=r, column=type_col+1).value}, Refund={val}")
        except ValueError:
            pass
print(f"Total refund rows found: {rows_found}")
wb.close()
