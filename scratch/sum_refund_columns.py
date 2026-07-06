import openpyxl

file_path = r'C:\Users\ahmad\Downloads\testing maret\income maret 2026.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

headers = [str(cell.value or '').strip() for cell in ws[1]]

refund_sums = {}
for r in range(2, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        h = headers[c-1]
        if 'pengembalian dana' in h.lower() or 'refund' in h.lower() or 'retur' in h.lower():
            val = ws.cell(row=r, column=c).value
            if val is not None:
                try:
                    refund_sums[h] = refund_sums.get(h, 0.0) + abs(float(val))
                except ValueError:
                    pass

print("Refund-related column sums in Keuangan file:")
for k, v in refund_sums.items():
    print(f"  {k}: {v:,.2f}")
wb.close()
