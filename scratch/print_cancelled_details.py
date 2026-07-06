import openpyxl

orders_file = r'C:\Users\ahmad\Downloads\testing maret\pesanan maret 2026.xlsx'
wb_ord = openpyxl.load_workbook(orders_file)
ws_ord = wb_ord.active
headers = [str(cell.value or '').strip() for cell in ws_ord[1]]
status_col = next((i for i, h in enumerate(headers) if 'status' in h.lower()), -1)

# Print row index and details of Dibatalkan orders where we show candidate sums
print("Cancelled orders in pesanan file:")
for r in range(2, ws_ord.max_row + 1):
    status = str(ws_ord.cell(row=r, column=status_col+1).value or '').strip()
    if status == 'Dibatalkan':
        print(f"  OrderID={ws_ord.cell(row=r, column=1).value}: "
              f"SubtotalAfterDisc={ws_ord.cell(row=r, column=16).value}, "
              f"SubtotalBeforeDisc={ws_ord.cell(row=r, column=13).value}, "
              f"OrderAmount={ws_ord.cell(row=r, column=29).value}")
wb_ord.close()
