import openpyxl

file_path = r'C:\Users\ahmad\Downloads\testing maret\income maret 2026.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

headers_inc = [str(cell.value or '').strip() for cell in ws[1]]
order_id_col = next((i for i, h in enumerate(headers_inc) if 'pesanan' in h.lower() or 'order' in h.lower()), -1)

# Load orders first to filter by March creation
orders_file = r'C:\Users\ahmad\Downloads\testing maret\pesanan maret 2026.xlsx'
wb_ord = openpyxl.load_workbook(orders_file)
ws_ord = wb_ord.active
headers_ord = [str(cell.value or '').strip().lower() for cell in ws_ord[1]]
ord_id_col = next((i for i, h in enumerate(headers_ord) if 'order id' in h or 'id pesanan' in h), -1)
date_col_ord = next((i for i, h in enumerate(headers_ord) if 'created' in h or 'waktu' in h or 'tanggal' in h), -1)

march_orders = set()
for r in range(2, ws_ord.max_row + 1):
    oid = str(ws_ord.cell(row=r, column=ord_id_col+1).value or '').strip()
    dt = str(ws_ord.cell(row=r, column=date_col_ord+1).value or '').strip()
    if '2026-03' in dt or '/03/2026' in dt or '03-2026' in dt:
        march_orders.add(oid)
wb_ord.close()

totals = {h: 0.0 for h in headers_inc if h}

for row_idx in range(2, ws.max_row + 1):
    oid = str(ws.cell(row=row_idx, column=order_id_col+1).value or '').strip()
    if oid not in march_orders:
        continue
    for c_idx, h in enumerate(headers_inc):
        if not h:
            continue
        val = ws.cell(row=row_idx, column=c_idx+1).value
        if val is not None:
            try:
                totals[h] += float(val)
            except ValueError:
                pass

print("Totals for March-created orders in Keuangan file:")
for k, v in totals.items():
    if abs(v) > 0.01:
        print(f"  {k}: {v:,.2f}")
wb.close()
