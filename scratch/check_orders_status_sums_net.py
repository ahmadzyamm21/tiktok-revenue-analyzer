import openpyxl

orders_file = r'C:\Users\ahmad\Downloads\testing maret\pesanan maret 2026.xlsx'
wb_ord = openpyxl.load_workbook(orders_file)
ws_ord = wb_ord.active
headers_ord = [str(cell.value or '').strip().lower() for cell in ws_ord[1]]
ord_id_col = next((i for i, h in enumerate(headers_ord) if 'order id' in h or 'id pesanan' in h), -1)
status_col = next((i for i, h in enumerate(headers_ord) if 'status' in h), -1)

# Let's find any column containing 'setelah diskon' or 'omset' or 'harga'
price_cols = [i for i, h in enumerate(headers_ord) if 'diskon' in h or 'omset' in h or 'harga' in h or 'pembayaran' in h]
print("Candidate price columns:")
for idx in price_cols:
    print(f"  [{idx}] {headers_ord[idx]}")

# Let's sum specifically using column 'subtotal setelah diskon penjual' (let's find it exactly)
net_price_col = next((i for i, h in enumerate(headers_ord) if 'subtotal setelah diskon penjual' in h or 'subtotal setelah diskon' in h), -1)
if net_price_col == -1:
    net_price_col = next((i for i, h in enumerate(headers_ord) if 'harga setelah diskon' in h or 'pembayaran' in h), -1)

print(f"Using column: {headers_ord[net_price_col]}")

status_sums = {}
for r in range(2, ws_ord.max_row + 1):
    status = str(ws_ord.cell(row=r, column=status_col+1).value or '').strip()
    val = ws_ord.cell(row=r, column=net_price_col+1).value
    price = 0.0
    if val is not None:
        try:
            price = float(val)
        except ValueError:
            pass
    status_sums[status] = status_sums.get(status, 0.0) + price

print("Sums by status:")
for k, v in status_sums.items():
    print(f"  {k}: {v:,.2f}")
wb_ord.close()
