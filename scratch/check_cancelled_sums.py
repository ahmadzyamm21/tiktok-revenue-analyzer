import openpyxl

orders_file = r'C:\Users\ahmad\Downloads\testing maret\pesanan maret 2026.xlsx'
wb_ord = openpyxl.load_workbook(orders_file)
ws_ord = wb_ord.active
headers = [str(cell.value or '').strip() for cell in ws_ord[1]]

status_col = next((i for i, h in enumerate(headers) if 'status' in h.lower()), -1)

# Let's find columns related to price or refund
for idx, h in enumerate(headers):
    hl = h.lower()
    if 'refund' in hl or 'amount' in hl or 'subtotal' in hl:
        # Let's sum this column specifically for 'Dibatalkan' status
        total_val = 0.0
        for r in range(2, ws_ord.max_row + 1):
            status = str(ws_ord.cell(row=r, column=status_col+1).value or '').strip()
            if status == 'Dibatalkan':
                val = ws_ord.cell(row=r, column=idx+1).value
                if val is not None:
                    try:
                        total_val += float(val)
                    except ValueError:
                        pass
        print(f"  {h} (col {idx}) sum for Dibatalkan: {total_val:,.2f}")
wb_ord.close()
