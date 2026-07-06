import openpyxl

orders_file = r'C:\Users\ahmad\Downloads\testing maret\pesanan maret 2026.xlsx'
wb_ord = openpyxl.load_workbook(orders_file)
ws_ord = wb_ord.active

status_sums_after_discount = {}
total_order_refund_amount = 0.0

for r in range(2, ws_ord.max_row + 1):
    status = str(ws_ord.cell(row=r, column=2).value or '').strip() # Order Status is col 1 (0-indexed: col 1)
    
    # Let's verify status by reading col 1 (Order Status)
    status_val = str(ws_ord.cell(row=r, column=2).value or '').strip()
    
    val_after_discount = ws_ord.cell(row=r, column=16).value # SKU Subtotal After Discount (col 15, 0-indexed: col 15+1=16)
    val_refund = ws_ord.cell(row=r, column=23).value # Order Refund Amount (col 22, 0-indexed: 23)
    
    p_after_disc = 0.0
    if val_after_discount is not None:
        try:
            p_after_disc = float(val_after_discount)
        except ValueError:
            pass
            
    p_ref = 0.0
    if val_refund is not None:
        try:
            p_ref = float(val_refund)
        except ValueError:
            pass
            
    status_sums_after_discount[status_val] = status_sums_after_discount.get(status_val, 0.0) + p_after_disc
    total_order_refund_amount += p_ref

print("Sums of SKU Subtotal After Discount by Order Status:")
for k, v in status_sums_after_discount.items():
    print(f"  {k}: {v:,.2f}")
print(f"Total Order Refund Amount (col 22): {total_order_refund_amount:,.2f}")
wb_ord.close()
