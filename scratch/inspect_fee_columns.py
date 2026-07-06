import openpyxl

file_path = r'C:\Users\ahmad\Downloads\testing maret\income maret 2026.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

headers = [str(cell.value or '').strip() for cell in ws[1]]

# Find "Total Biaya" column and all "Biaya" columns
biaya_cols = []
total_biaya_col = -1
for i, h in enumerate(headers):
    if 'biaya' in h.lower() or 'komisi' in h.lower() or 'fee' in h.lower():
        biaya_cols.append(i)
        if h.lower().startswith('total biaya'):
            total_biaya_col = i

print(f"Total Biaya column: [{total_biaya_col}] {headers[total_biaya_col] if total_biaya_col >= 0 else 'NOT FOUND'}")
print(f"\nAll fee-related columns:")
for idx in biaya_cols:
    # Sum column values
    total = 0.0
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=idx+1).value
        if val is not None:
            try:
                total += float(val)
            except ValueError:
                pass
    print(f"  [{idx}] {headers[idx]}: {total:,.2f}")

wb.close()
