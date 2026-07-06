import openpyxl

file_path = r'C:\Users\ahmad\Downloads\testing maret\income maret 2026.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

headers = [str(cell.value or '').strip() for cell in ws[1]]
cols_to_sum = [9, 20, 21, 31, 57, 69, 71, 73, 75]

sums = {idx: 0.0 for idx in cols_to_sum}

for row_idx in range(2, ws.max_row + 1):
    for idx in cols_to_sum:
        val = ws.cell(row=row_idx, column=idx+1).value
        if val is not None:
            try:
                sums[idx] += abs(float(val))
            except ValueError:
                pass

print("Sum of discount-related columns (absolute values):")
for idx, s in sums.items():
    print(f"  {headers[idx]} (col {idx}): {s:,.2f}")
wb.close()
