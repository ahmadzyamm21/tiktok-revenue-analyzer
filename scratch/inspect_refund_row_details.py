import openpyxl

file_path = r'C:\Users\ahmad\Downloads\testing maret\income maret 2026.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

headers = [str(cell.value or '').strip() for cell in ws[1]]

# Print all row data for a few rows that have non-zero refund column values
refund_col = next((i for i, h in enumerate(headers) if 'pengembalian dana' in h.lower() or 'refund' in h.lower()), -1)
print(f"Refund col is: {headers[refund_col]} at col {refund_col}")

rows_printed = 0
for r in range(2, ws.max_row + 1):
    val = ws.cell(row=r, column=refund_col+1).value
    if val is not None:
        try:
            fval = float(val)
            if abs(fval) > 0.01:
                print(f"\nRow {r}:")
                for c in range(1, ws.max_column + 1):
                    cell_val = ws.cell(row=r, column=c).value
                    if cell_val is not None and str(cell_val).strip() != '':
                        print(f"  {headers[c-1]}: {cell_val}")
                rows_printed += 1
                if rows_printed >= 3:
                    break
        except ValueError:
            pass

wb.close()
